import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series


class ExcelExporter:
    def __init__(self, output_directory):
        self.output_directory = output_directory
        self.results_json_path = os.path.join(
            self.output_directory, "info", "results.json"
        )
        if not os.path.isfile(self.results_json_path):
            raise FileNotFoundError(
                f"El archivo JSON no existe en la ruta: {self.results_json_path}"
            )

    def process_json_to_excel(self):
        try:
            with open(self.results_json_path, "r") as f:
                data = json.load(f)

            excel_file_path = os.path.join(self.output_directory, "results.xlsx")
            with pd.ExcelWriter(excel_file_path, engine="openpyxl") as writer:
                self._write_ref_sheet(writer, data)
                for sample_name, sample_data in data.get("samples", {}).items():
                    self._write_sample_sheet(writer, sample_name, sample_data)

            print(f"Archivo Excel guardado en: {excel_file_path}")
            return excel_file_path
        except Exception as e:
            print(f"Error procesando el archivo JSON: {e}")
            raise

    def _write_ref_sheet(self, writer, data):
        ref_data = data.get("reference", {})
        metadata = data.get("metadata", {})
        scale_info = ref_data.get("scale", {})
        ref_df = pd.DataFrame(
            {
                "Property": [
                    "Scale Value",
                    "Scale Unit",
                    "Image Path",
                    "Created At",
                    "Description",
                ],
                "Value": [
                    scale_info.get("value", "N/A"),
                    scale_info.get("unit", "N/A"),
                    ref_data.get("image_path", "N/A"),
                    metadata.get("created_at", "N/A"),
                    metadata.get("description", "N/A"),
                ],
            }
        )
        ref_df.to_excel(writer, index=False, sheet_name="ref")

    def _write_sample_sheet(self, writer, sample_name, sample_data):
        base_data = {
            "Property": [key for key in sample_data.keys() if key != "distances"],
            "Value": [
                value for key, value in sample_data.items() if key != "distances"
            ],
        }
        sample_df = pd.DataFrame(base_data)
        sample_df.to_excel(writer, index=False, sheet_name=sample_name, startcol=0)

        distances = sample_data.get("distances", [])
        if distances:
            distance_rows = []
            for dist in distances:
                pair = dist.get("pair", {})
                keys = list(pair.keys())
                if len(keys) == 2:
                    id_1, id_2 = keys
                    x_1, y_1 = pair[id_1]
                    x_2, y_2 = pair[id_2]
                    distance = dist.get("distance", 0)
                    distance_rows.append([id_1, id_2, x_1, y_1, x_2, y_2, distance])
            distance_df = pd.DataFrame(
                distance_rows,
                columns=["id_1", "id_2", "x_1", "y_1", "x_2", "y_2", "distance"],
            )
            distance_df.to_excel(
                writer, index=False, sheet_name=sample_name, startcol=4, startrow=0
            )

    def _add_charts_to_samples(self, excel_file_path, data):
        wb = load_workbook(excel_file_path)
        for sample_name in data.get("samples", {}).keys():
            ws = wb[sample_name]
            chart = ScatterChart()
            chart.title = f"Scatter Plot for {sample_name}"
            chart.x_axis.title = "X Coordinate"
            chart.y_axis.title = "Y Coordinate"

            x_values = Reference(ws, min_col=6, min_row=2, max_row=ws.max_row)
            y_values = Reference(ws, min_col=7, min_row=2, max_row=ws.max_row)

            series = Series(y_values, x_values, title="Particle Pairs")
            chart.series.append(series)

            ws.add_chart(chart, "J1")  # Coloca el gráfico en una posición fija
        wb.save(excel_file_path)
